"""
Management command: seed_records

Reads ILIGAN_NAP-GRDS_Monitoring_Batching_Labeling_System_2025_COPY.xlsx and
populates Section, GRDSItem, and ArchiveBatch. See records/models.py for the
sheet -> model mapping.

INSTALL
-------
Put this file at:  records/management/commands/seed_records.py
(create the two `management/` and `management/commands/` folders, each with
an empty __init__.py, if they don't exist yet)

Update the "# >>> EDIT" import below if your app isn't called `records`,
then run:

    python manage.py seed_records "C:\\path\\to\\ILIGAN_NAP-GRDS_Monitoring_Batching_Labeling_System_2025_COPY.xlsx" --dry-run

WHAT GETS IMPORTED
-------------------
- REFERENCE sheet (columns A-C only)       -> GRDSItem catalog
- REFERENCE sheet (column E)               -> Section names
- BRANCH COMMONLY USED ITEMS sheet         -> flags GRDSItem.is_commonly_used
- DATA ENTRY FOR DISPOSAL sheet            -> ArchiveBatch (batch_type=disposal)
- DATA ENTRY ARCHIVING sheet               -> ArchiveBatch (batch_type=archiving)
- A second pass links disposal <-> archiving batches together via the
  'BATCH NO FROM ARCHIVES' / 'STATUS BATCH NO DISPOSAL' columns.

NOT imported: the two LABEL FORM sheets and the NAP Form 3 sheets. Those are
printable outputs/official forms generated FROM this data, plus REFERENCE
columns F/K/L/M which are just a leftover running tally of allocated batch
numbers, not real per-item data. DisposalRequest (the NAP Form 3 model) is
left for you to fill in through the admin when you actually file one — the
form's line items don't carry enough detail to auto-match specific boxes.

QUIRKS HANDLED
--------------
- GRDS item numbers are a mix of plain numbers ('26', '74') and prefixed
  codes ('SSSRDS 9'); item_no is stored as text throughout.
- Retention periods, sections, and locations are frequently blank, richly
  formatted text, or inconsistent — all treated as optional/free text.
- Rows with no batch_no are skipped (nothing to key off of); everything
  else is imported even if some fields are blank.
"""

import datetime
import decimal
import re

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone as dj_timezone

# >>> EDIT: point this at your real app's models module if different
from records.models import Section, GRDSItem, ArchiveBatch  # noqa


def norm(s):
    if s is None:
        return ''
    return re.sub(r'\s+', ' ', str(s)).strip()


def to_int(value):
    if value is None:
        return None
    if isinstance(value, (int, float, decimal.Decimal)):
        return int(value)
    text = norm(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def to_date(value):
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    return None


SCANNING_MAP = {
    'scanned': GRDSItem.ScanningStatus.SCANNED,
    'unscanned': GRDSItem.ScanningStatus.UNSCANNED,
    'not applicable': GRDSItem.ScanningStatus.NOT_APPLICABLE,
}


class Command(BaseCommand):
    help = "Seed the records app (catalog + archive/disposal batches) from the GRDS Excel workbook."

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', type=str)
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        path = options['xlsx_path']
        dry_run = options['dry_run']

        self.stdout.write(f"Loading {path} ...")
        wb = openpyxl.load_workbook(path, data_only=True)

        with transaction.atomic():
            self.grds_by_item_no = {}

            sec_count = self._seed_sections(wb['REFERENCE'])
            self.stdout.write(self.style.SUCCESS(f"Sections: {sec_count}"))

            cat_count = self._seed_catalog(wb['REFERENCE'])
            self.stdout.write(self.style.SUCCESS(f"GRDSItems from REFERENCE: {cat_count}"))

            common_count = self._seed_common_items(wb['BRANCH COMMONLY USED ITEMS'])
            self.stdout.write(self.style.SUCCESS(
                f"GRDSItems from BRANCH COMMONLY USED ITEMS: {common_count} (flagged/created)"))

            arch_count, arch_skipped = self._seed_batches(
                wb['DATA ENTRY ARCHIVING'], ArchiveBatch.BatchType.ARCHIVING, has_location=False)
            self.stdout.write(self.style.SUCCESS(
                f"Archiving batches: {arch_count} ({arch_skipped} rows skipped, no batch no)"))

            disp_count, disp_skipped = self._seed_batches(
                wb['DATA ENTRY FOR DISPOSAL'], ArchiveBatch.BatchType.DISPOSAL, has_location=True)
            self.stdout.write(self.style.SUCCESS(
                f"Disposal batches: {disp_count} ({disp_skipped} rows skipped, no batch no)"))

            link_count = self._link_batches(wb)
            self.stdout.write(self.style.SUCCESS(f"Linked archiving<->disposal pairs: {link_count}"))

            if dry_run:
                self.stdout.write(self.style.WARNING("--dry-run set: rolling back all changes."))
                transaction.set_rollback(True)
            else:
                self.stdout.write(self.style.SUCCESS("Done. Changes committed."))

    # ------------------------------------------------------------------

    def _seed_sections(self, ws):
        seen = set()
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = norm(row[4]) if len(row) > 4 else ''
            if not name or name in seen:
                continue
            seen.add(name)
            Section.objects.get_or_create(name=name)
            count += 1
        return count

    def _get_or_create_section(self, name):
        name = norm(name)
        if not name:
            return None
        section, _ = Section.objects.get_or_create(name=name)
        return section

    # ------------------------------------------------------------------

    def _seed_catalog(self, ws):
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            item_no, title, retention = row[0], row[1], row[2]
            item_no = norm(item_no) if item_no is not None else ''
            title = norm(title)
            if not item_no or not title:
                continue
            grds_item, _ = GRDSItem.objects.get_or_create(
                item_no=item_no,
                defaults=dict(records_series_title=title[:255], retention_period=norm(retention)[:150]),
            )
            self.grds_by_item_no[item_no] = grds_item
            count += 1
        return count

    def _seed_common_items(self, ws):
        count = 0
        for row in ws.iter_rows(min_row=3, values_only=True):  # row 1 = banner, row 2 = header
            item_no, title, retention = row[0], row[1], row[2]
            item_no = norm(item_no) if item_no is not None else ''
            title = norm(title)
            if not item_no or not title:
                continue
            grds_item, _ = GRDSItem.objects.get_or_create(
                item_no=item_no,
                defaults=dict(records_series_title=title[:255], retention_period=norm(retention)[:150]),
            )
            if not grds_item.is_commonly_used:
                grds_item.is_commonly_used = True
                grds_item.save(update_fields=['is_commonly_used'])
            self.grds_by_item_no[item_no] = grds_item
            count += 1
        return count

    def _resolve_grds_item(self, item_no_raw):
        item_no = norm(item_no_raw)
        if not item_no:
            return None
        return self.grds_by_item_no.get(item_no)

    # ------------------------------------------------------------------
    # Shared column layout for DATA ENTRY ARCHIVING / DATA ENTRY FOR DISPOSAL:
    # 0 ref_no, 1 box, 2 batch_no, 3 date, 4 grds_item_no, 5 title (unused,
    # GRDSItem already carries it), 6 section, 7 description, 8 period,
    # 9 latest_year, 10 years_as_of, 11 retention_years, 12 status_value,
    # 13 scanning, 14 cross-link batch no, 15 location (disposal sheet only)
    # ------------------------------------------------------------------

    def _seed_batches(self, ws, batch_type, has_location):
        created, skipped = 0, 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            batch_no = norm(row[2])
            if not batch_no:
                skipped += 1
                continue

            grds_item_no_raw = norm(row[4])
            grds_item = self._resolve_grds_item(grds_item_no_raw)
            section = self._get_or_create_section(row[6])
            scanning_raw = norm(row[13]).lower()
            scanning_status = SCANNING_MAP.get(scanning_raw, '')

            defaults = dict(
                batch_type=batch_type,
                reference_no=norm(row[0]),
                box_number=norm(row[1]),
                batch_date=to_date(row[3]) or dj_timezone.now().date(),
                grds_item=grds_item,
                grds_item_no_raw=grds_item_no_raw,
                section=section,
                description=norm(row[7])[:255],
                period_covered=norm(row[8])[:255],
                latest_year=to_int(row[9]),
                years_as_of_count=to_int(row[10]),
                retention_period_years=to_int(row[11]),
                disposal_status_value=to_int(row[12]),
                scanning_status=scanning_status,
            )
            if has_location:
                defaults['location'] = norm(row[15])[:150]

            ArchiveBatch.objects.get_or_create(batch_no=batch_no, defaults=defaults)
            created += 1
        return created, skipped

    # ------------------------------------------------------------------

    def _link_batches(self, wb):
        """Second pass: connect a disposal batch back to the archiving batch
        it came from, using the cross-reference column on each sheet."""
        linked = 0

        for row in wb['DATA ENTRY FOR DISPOSAL'].iter_rows(min_row=2, values_only=True):
            batch_no, cross_ref = norm(row[2]), norm(row[14])
            if not batch_no or not cross_ref:
                continue
            try:
                disposal = ArchiveBatch.objects.get(batch_no=batch_no)
                archiving = ArchiveBatch.objects.get(batch_no=cross_ref)
            except ArchiveBatch.DoesNotExist:
                continue
            if disposal.linked_batch_id != archiving.pk:
                disposal.linked_batch = archiving
                disposal.save(update_fields=['linked_batch'])
                linked += 1

        for row in wb['DATA ENTRY ARCHIVING'].iter_rows(min_row=2, values_only=True):
            batch_no, cross_ref = norm(row[2]), norm(row[14])
            if not batch_no or not cross_ref:
                continue
            try:
                archiving = ArchiveBatch.objects.get(batch_no=batch_no)
                disposal = ArchiveBatch.objects.get(batch_no=cross_ref)
            except ArchiveBatch.DoesNotExist:
                continue
            if archiving.linked_batch_id != disposal.pk:
                archiving.linked_batch = disposal
                archiving.save(update_fields=['linked_batch'])
                linked += 1

        return linked