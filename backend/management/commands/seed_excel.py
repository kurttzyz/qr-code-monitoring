"""
Management command: seed_excel

Reads DBASE_SUPPLIES_REQUEST___INVENTORY_2025_COPY.xlsx and populates the
Django models (Category, Unit, Product, ItemRequest/Item, Receiving/Item,
StockRelease/Item) so you don't have to encode everything by hand.

INSTALL
-------
Put this file at:  <your_app>/management/commands/seed_excel.py
(create the two `management/` and `management/commands/` folders, each with
an empty __init__.py, if they don't exist yet)

Update the two lines marked "# >>> EDIT" below to match your actual app name
and the location of the xlsx file, then run:

    python manage.py seed_excel /path/to/DBASE_SUPPLIES_REQUEST___INVENTORY_2025_COPY.xlsx

The command is idempotent-ish: re-running it will NOT duplicate Products
(matched by SKU/description) or Requests/Receivings (matched by reference
number), but it WILL add new StockRelease rows each run since the source
sheet has no natural per-row reference number. Delete existing StockRelease
rows first if you need to re-run that part cleanly.

WHAT GETS IMPORTED
-------------------
- REFERENCE sheet              -> Category, Unit, Product (the master catalog)
- DATA ENTRY - REQUEST sheet   -> ItemRequest + ItemRequestItem
- DATA ENTRY - RECEIVED sheet  -> Receiving + ReceivingItem
- DATA ENTRY - RELEASED sheet  -> StockRelease + StockReleaseItem

NOT imported (too inconsistent / out of scope for a first pass): the janitorial
sheet, the "SUMMARY"/"SUPPLIES SUMMARY" roll-up sheets (these are computed
views of the same data, not source data), and the old per-month 2021-2026
snapshot tabs. Ask for a follow-up pass if you want those too.

ASSUMPTIONS (documented so you can change them)
------------------------------------------------
- All stock is assumed to live in one Warehouse ("MAIN") / one
  WarehouseLocation ("MAIN-01"), since the workbook doesn't track locations.
- Receiving rows have no supplier column in the sheet, so they're all
  attached to a placeholder Supplier ("UNSPECIFIED"). Edit it in the admin
  afterwards if you know the real suppliers.
- ItemRequest.requested_by / Receiving.received_by / StockRelease.released_by
  all point at a single placeholder "excel_import" system user (created if
  missing), since the sheet doesn't name individual staff. StockRelease's
  *department/recipient* field does carry the real "REQUESTED BY" values
  from the sheet.
- Product SKU: uses the sheet's numeric "Mat Code" when present and not
  already used for a different description; otherwise a generated code
  (GEN-000001, GEN-000002, ...) is assigned and kept stable via a
  description-based cache.
"""

import argparse
import datetime
import decimal
import re
import sys
from collections import defaultdict

import openpyxl
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone as dj_timezone

# >>> EDIT: point this at your real app's models module if different
from backend.models import (  # noqa
    Category, Unit, Product, Supplier, Warehouse, WarehouseLocation,
    ItemRequest, ItemRequestItem, Receiving, ReceivingItem,
    StockRelease, StockReleaseItem,
)

User = get_user_model()


def norm(s):
    if s is None:
        return ''
    return re.sub(r'\s+', ' ', str(s)).strip()


def norm_key(s):
    return norm(s).lower()


def to_qty(value):
    """Coerce a cell value to a Decimal quantity. Blank/dash cells are a
    normal zero. Cells with real text in them that isn't a number (e.g.
    'NONE IN STOCK') are also treated as zero, but the original text is
    returned as `issue` so the caller can flag it instead of dropping the row."""
    if value is None:
        return decimal.Decimal('0'), None
    if isinstance(value, (int, float, decimal.Decimal)):
        return decimal.Decimal(str(value)), None
    text = norm(value)
    if not text or text == '-':
        return decimal.Decimal('0'), None
    try:
        return decimal.Decimal(text), None
    except (decimal.InvalidOperation, ValueError):
        return decimal.Decimal('0'), text


def to_aware(value):
    """Turn an Excel datetime cell into a timezone-aware datetime Django is
    happy with, or None if it isn't a real date."""
    if not isinstance(value, (datetime.date, datetime.datetime)):
        return None
    if isinstance(value, datetime.date) and not isinstance(value, datetime.datetime):
        value = datetime.datetime(value.year, value.month, value.day)
    if dj_timezone.is_naive(value):
        value = dj_timezone.make_aware(value, dj_timezone.get_default_timezone())
    return value


class Command(BaseCommand):
    help = "Seed catalog + request/receiving/release history from the supplies Excel workbook."

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', type=str)
        parser.add_argument('--dry-run', action='store_true',
                             help="Parse and report counts without writing to the database.")

    def handle(self, *args, **options):
        path = options['xlsx_path']
        dry_run = options['dry_run']

        self.stdout.write(f"Loading {path} ...")
        wb = openpyxl.load_workbook(path, data_only=True)

        with transaction.atomic():
            self.system_user = self._get_or_create_system_user()
            self.warehouse, self.location = self._get_or_create_default_location()
            self.default_supplier = self._get_or_create_default_supplier()

            self.product_by_sku = {}          # sku -> Product
            self.product_by_desc = {}         # normalized description -> Product
            self._gen_code_counter = 0

            cat_count, unit_count, prod_count = self._seed_catalog(wb['REFERENCE'])
            self.stdout.write(self.style.SUCCESS(
                f"Catalog: {cat_count} categories, {unit_count} units, {prod_count} products"))

            req_count, req_item_count, req_skipped = self._seed_requests(wb['DATA ENTRY - REQUEST'])
            self.stdout.write(self.style.SUCCESS(
                f"Requests: {req_count} ItemRequests, {req_item_count} items ({req_skipped} rows skipped)"))

            rcv_count, rcv_item_count, rcv_skipped = self._seed_receiving(wb['DATA ENTRY - RECEIVED'])
            self.stdout.write(self.style.SUCCESS(
                f"Receiving: {rcv_count} Receivings, {rcv_item_count} items ({rcv_skipped} rows skipped)"))

            rel_count, rel_skipped = self._seed_releases(wb['DATA ENTRY - RELEASED'])
            self.stdout.write(self.style.SUCCESS(
                f"Releases: {rel_count} StockReleases created ({rel_skipped} rows skipped)"))

            if dry_run:
                self.stdout.write(self.style.WARNING("--dry-run set: rolling back all changes."))
                transaction.set_rollback(True)
            else:
                self.stdout.write(self.style.SUCCESS("Done. Changes committed."))

    # ------------------------------------------------------------------
    # Bootstrapping helpers
    # ------------------------------------------------------------------

    def _get_or_create_system_user(self):
        user, _ = User.objects.get_or_create(
            username='excel_import',
            defaults=dict(first_name='Excel', last_name='Import',
                           role=getattr(User.Role, 'ADMIN', 'admin'),
                           is_active_employee=True),
        )
        return user

    def _get_or_create_default_location(self):
        warehouse, _ = Warehouse.objects.get_or_create(
            code='MAIN', defaults=dict(name='Main Warehouse'))
        location, _ = WarehouseLocation.objects.get_or_create(
            warehouse=warehouse, code='MAIN-01',
            defaults=dict(name='Main Storage', location_type='zone'))
        return warehouse, location

    def _get_or_create_default_supplier(self):
        supplier, _ = Supplier.objects.get_or_create(
            code='UNSPEC', defaults=dict(company_name='Unspecified Supplier (from Excel import)'))
        return supplier

    def _get_or_create_unit(self, unit_text, cache):
        key = norm_key(unit_text) or 'piece'
        if key in cache:
            return cache[key]
        display = norm(unit_text) or 'Piece'
        unit, _ = Unit.objects.get_or_create(
            name=display, defaults=dict(abbreviation=display[:10]))
        cache[key] = unit
        return unit

    def _get_or_create_category(self, cat_text, cache):
        key = norm_key(cat_text) or 'uncategorized'
        if key in cache:
            return cache[key]
        display = norm(cat_text) or 'Uncategorized'
        category, _ = Category.objects.get_or_create(name=display)
        cache[key] = category
        return category

    def _next_generated_sku(self):
        self._gen_code_counter += 1
        return f"GEN-{self._gen_code_counter:06d}"

    # ------------------------------------------------------------------
    # REFERENCE -> Category / Unit / Product
    # ------------------------------------------------------------------

    def _seed_catalog(self, ws):
        unit_cache, cat_cache = {}, {}
        categories_seen, units_seen = set(), set()

        for row in ws.iter_rows(min_row=1, values_only=True):
            mat_code, description, category_text, unit_text = row[0], row[1], row[2], row[3]

            if norm(mat_code) == 'Mat Code':
                continue  # section header row
            description = norm(description)
            if not description:
                continue

            category = self._get_or_create_category(category_text, cat_cache)
            unit = self._get_or_create_unit(unit_text, unit_cache)
            categories_seen.add(category.pk)
            units_seen.add(unit.pk)

            desc_key = norm_key(description)
            if desc_key in self.product_by_desc:
                continue  # already created from an earlier row with the same description

            sku = None
            if mat_code not in (None, 0, '0', '-'):
                candidate = str(mat_code)
                if candidate not in self.product_by_sku:
                    sku = candidate
            if sku is None:
                sku = self._next_generated_sku()

            product, _ = Product.objects.get_or_create(
                sku=sku,
                defaults=dict(name=description[:200], category=category, unit=unit),
            )
            self.product_by_sku[sku] = product
            self.product_by_desc[desc_key] = product

        return len(categories_seen), len(units_seen), len(self.product_by_desc)

    def _resolve_product(self, description, unit_text, category_text=('Uncategorized',)):
        """Find a product by description, creating a fallback one if the
        REFERENCE sheet didn't already define it (some request/release rows
        mention items missing from REFERENCE)."""
        description = norm(description)
        if not description:
            return None
        desc_key = norm_key(description)
        product = self.product_by_desc.get(desc_key)
        if product:
            return product

        unit, _ = Unit.objects.get_or_create(
            name=norm(unit_text) or 'Piece', defaults=dict(abbreviation=(norm(unit_text) or 'Piece')[:10]))
        category, _ = Category.objects.get_or_create(name='Uncategorized')
        sku = self._next_generated_sku()
        product = Product.objects.create(sku=sku, name=description[:200], category=category, unit=unit)
        self.product_by_sku[sku] = product
        self.product_by_desc[desc_key] = product
        return product

    # ------------------------------------------------------------------
    # DATA ENTRY - REQUEST -> ItemRequest / ItemRequestItem
    # ------------------------------------------------------------------

    def _seed_requests(self, ws):
        rows_by_ris = defaultdict(list)
        for row in ws.iter_rows(min_row=2, values_only=True):
            ris_no = norm(row[0])
            if not ris_no:
                continue
            rows_by_ris[ris_no].append(row)

        created, items, skipped = 0, 0, 0
        for ris_no, rows in rows_by_ris.items():
            first = rows[0]
            request_date = to_aware(first[3])

            item_request, was_created = ItemRequest.objects.get_or_create(
                reference_number=ris_no,
                defaults=dict(department='UNSPECIFIED', requested_by=self.system_user,
                              purpose='Imported from Excel'),
            )
            if was_created and request_date:
                ItemRequest.objects.filter(pk=item_request.pk).update(created_at=request_date)
            if was_created:
                created += 1

            issues = []
            for row in rows:
                description, unit_text = row[4], row[6]
                if not norm(description):
                    skipped += 1  # nothing to attach the row to at all
                    continue
                qty, issue = to_qty(row[5])
                product = self._resolve_product(description, unit_text)
                if not product:
                    skipped += 1
                    continue
                ItemRequestItem.objects.get_or_create(
                    request=item_request, product=product,
                    defaults=dict(quantity_requested=qty),
                )
                items += 1
                if issue:
                    issues.append(f"{product.name}: qty cell was '{issue}', set to 0")

            if was_created and issues:
                note = "Imported from Excel — flagged rows: " + "; ".join(issues)
                ItemRequest.objects.filter(pk=item_request.pk).update(purpose=note[:2000])

        return created, items, skipped

    # ------------------------------------------------------------------
    # DATA ENTRY - RECEIVED -> Receiving / ReceivingItem
    # ------------------------------------------------------------------

    def _seed_receiving(self, ws):
        rows_by_ris = defaultdict(list)
        for row in ws.iter_rows(min_row=2, values_only=True):
            ris_no = norm(row[0])
            if not ris_no:
                continue
            rows_by_ris[ris_no].append(row)

        created, items, skipped = 0, 0, 0
        for ris_no, rows in rows_by_ris.items():
            ref = f"RCV-{ris_no}"
            first = rows[0]
            received_date = to_aware(first[3])

            receiving, was_created = Receiving.objects.get_or_create(
                reference_number=ref,
                defaults=dict(supplier=self.default_supplier, warehouse=self.warehouse,
                              status='stored', received_by=self.system_user,
                              remarks='Imported from Excel'),
            )
            if was_created and received_date:
                Receiving.objects.filter(pk=receiving.pk).update(created_at=received_date)
            if was_created:
                created += 1

            issues = []
            for row in rows:
                description, unit_text = row[4], row[6]
                if not norm(description):
                    skipped += 1  # nothing to attach the row to at all
                    continue
                qty, issue = to_qty(row[5])
                product = self._resolve_product(description, unit_text)
                if not product:
                    skipped += 1
                    continue
                ReceivingItem.objects.get_or_create(
                    receiving=receiving, product=product, location=self.location,
                    defaults=dict(quantity_received=qty, quantity_accepted=qty, is_stored=True),
                )
                items += 1
                if issue:
                    issues.append(f"{product.name}: qty cell was '{issue}', set to 0")

            if was_created and issues:
                note = "Imported from Excel — flagged rows: " + "; ".join(issues)
                Receiving.objects.filter(pk=receiving.pk).update(remarks=note[:2000])

        return created, items, skipped

    # ------------------------------------------------------------------
    # DATA ENTRY - RELEASED -> StockRelease / StockReleaseItem
    # ------------------------------------------------------------------

    def _seed_releases(self, ws):
        created, skipped = 0, 0
        seq = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            release_date_raw, description, qty_raw, unit_text, stock_code, dept = row[0], row[1], row[2], row[3], row[4], row[5]
            if not norm(description):
                skipped += 1  # nothing to attach the row to at all
                continue
            qty, issue = to_qty(qty_raw)

            product = self._resolve_product(description, unit_text)
            if not product:
                skipped += 1
                continue

            seq += 1
            ref = f"REL-{seq:06d}"
            department = norm(dept) or 'UNSPECIFIED'
            remarks = 'Imported from Excel'
            if issue:
                remarks += f" — flagged: qty cell was '{issue}', set to 0"

            release = StockRelease.objects.create(
                reference_number=ref, location=self.location, released_by=self.system_user,
                released_to_department=department, recipient_name=department,
                status='released', remarks=remarks,
            )
            release_date = to_aware(release_date_raw)
            if release_date:
                StockRelease.objects.filter(pk=release.pk).update(created_at=release_date)

            StockReleaseItem.objects.create(release=release, product=product, quantity=qty)
            created += 1

        return created, skipped